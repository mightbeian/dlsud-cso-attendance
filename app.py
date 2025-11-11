from flask import Flask, render_template, request, jsonify, send_file
from database import db, User, Attendance
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///attendance.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

with app.app_context():
    db.create_all()
    # Add sample users if database is empty
    if User.query.count() == 0:
        sample_users = [
            User(id_number='20212345', full_name='Juan Dela Cruz', birthday='01-15'),
            User(id_number='20212346', full_name='Maria Santos', birthday='03-22'),
            User(id_number='20212347', full_name='Pedro Reyes', birthday='07-08')
        ]
        for user in sample_users:
            db.session.add(user)
        db.session.commit()
        print("Sample users added to database.")

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/check-status', methods=['POST'])
def check_status():
    """Check if user is currently timed in"""
    data = request.json
    id_number = data.get('id_number', '').strip()
    
    if not id_number:
        return jsonify({'error': 'ID Number is required'}), 400
    
    user = User.query.filter_by(id_number=id_number).first()
    if not user:
        return jsonify({'error': 'ID Number not found in the system'}), 404
    
    # Check latest attendance record
    latest = Attendance.query.filter_by(user_id=user.id).order_by(Attendance.timestamp.desc()).first()
    
    is_timed_in = latest and latest.event_type == 'In'
    
    return jsonify({
        'user_id': user.id,
        'full_name': user.full_name,
        'is_timed_in': is_timed_in
    })

@app.route('/api/time-in', methods=['POST'])
def time_in():
    data = request.json
    id_number = data.get('id_number', '').strip()
    
    if not id_number:
        return jsonify({'error': 'ID Number is required'}), 400
    
    user = User.query.filter_by(id_number=id_number).first()
    if not user:
        return jsonify({'error': 'ID Number not found in the system'}), 404
    
    # Check if already timed in
    latest = Attendance.query.filter_by(user_id=user.id).order_by(Attendance.timestamp.desc()).first()
    if latest and latest.event_type == 'In':
        return jsonify({'error': f'{user.full_name}, you are already timed in'}), 400
    
    # Create new time in record
    attendance = Attendance(user_id=user.id, event_type='In')
    db.session.add(attendance)
    db.session.commit()
    
    # Check for birthday
    today = datetime.now()
    current_month_day = today.strftime('%m-%d')
    is_birthday = user.birthday == current_month_day
    
    message = f"Happy Birthday, {user.full_name}! You are timed in." if is_birthday else f"Welcome, {user.full_name}! You are timed in."
    
    return jsonify({
        'success': True,
        'message': message,
        'is_birthday': is_birthday,
        'timestamp': attendance.timestamp.strftime('%Y-%m-%d %I:%M:%S %p')
    })

@app.route('/api/time-out', methods=['POST'])
def time_out():
    data = request.json
    id_number = data.get('id_number', '').strip()
    
    if not id_number:
        return jsonify({'error': 'ID Number is required'}), 400
    
    user = User.query.filter_by(id_number=id_number).first()
    if not user:
        return jsonify({'error': 'ID Number not found in the system'}), 404
    
    # Check if user is timed in
    latest = Attendance.query.filter_by(user_id=user.id).order_by(Attendance.timestamp.desc()).first()
    if not latest or latest.event_type == 'Out':
        return jsonify({'error': f'{user.full_name}, you are not currently timed in'}), 400
    
    # Create new time out record
    attendance = Attendance(user_id=user.id, event_type='Out')
    db.session.add(attendance)
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': f"Goodbye, {user.full_name}! You are timed out.",
        'timestamp': attendance.timestamp.strftime('%Y-%m-%d %I:%M:%S %p')
    })

@app.route('/api/export-dtr')
def export_dtr():
    """Export attendance logs as Excel DTR format"""
    # Get all attendance records grouped by user and date
    attendance_records = db.session.query(
        User.full_name,
        Attendance.timestamp,
        Attendance.event_type
    ).join(User).order_by(Attendance.timestamp.desc()).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "DTR Report"
    
    # Style definitions
    header_fill = PatternFill(start_color="00693C", end_color="00693C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = ['Date', 'Full Name', 'Time In', 'Time Out', 'Total Hours']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Process records
    dtr_data = {}
    for name, timestamp, event_type in attendance_records:
        date_key = timestamp.strftime('%Y-%m-%d')
        user_date_key = f"{name}_{date_key}"
        
        if user_date_key not in dtr_data:
            dtr_data[user_date_key] = {
                'date': date_key,
                'name': name,
                'time_in': None,
                'time_out': None
            }
        
        if event_type == 'In':
            if dtr_data[user_date_key]['time_in'] is None:
                dtr_data[user_date_key]['time_in'] = timestamp
        else:
            if dtr_data[user_date_key]['time_out'] is None:
                dtr_data[user_date_key]['time_out'] = timestamp
    
    # Write data rows
    row = 2
    for key in sorted(dtr_data.keys(), reverse=True):
        record = dtr_data[key]
        
        ws.cell(row=row, column=1, value=record['date'])
        ws.cell(row=row, column=2, value=record['name'])
        ws.cell(row=row, column=3, value=record['time_in'].strftime('%I:%M %p') if record['time_in'] else '')
        ws.cell(row=row, column=4, value=record['time_out'].strftime('%I:%M %p') if record['time_out'] else '')
        
        # Calculate total hours
        if record['time_in'] and record['time_out']:
            delta = record['time_out'] - record['time_in']
            hours = delta.total_seconds() / 3600
            ws.cell(row=row, column=5, value=f"{hours:.2f}")
        else:
            ws.cell(row=row, column=5, value='')
        
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    
    # Save file
    filename = f"DTR_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join('exports', filename)
    
    # Create exports directory if it doesn't exist
    os.makedirs('exports', exist_ok=True)
    
    wb.save(filepath)
    
    return send_file(filepath, as_attachment=True, download_name=filename)

if __name__ == '__main__':
    app.run(debug=False, host='127.0.0.1', port=5000)